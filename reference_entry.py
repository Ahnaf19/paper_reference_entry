from openpyxl import load_workbook

xlsx_file_directory = r''
wb = load_workbook(xlsx_file_directory)  
ws = wb["Sheet1"]  # Work Sheet
citation_column = ws['C']  # Citation Column
citation_column_list = [citation_column[x].value for x in range(len(citation_column))]

# print(len(citation_column_list))

citation_entry = input("Paste Harvard citation: ")
if citation_entry not in citation_column_list:
    RefSL_entry = len(citation_column)
    p_w_entry = input('p/w: ')
    year_entry = input("Year: ")
    link_entry = input("DOI/link: ")
    new_entry = {
      'Ref SL': RefSL_entry,
      'p/w': p_w_entry 
      'Citation': citation_entry,
      'Year': year_entry,
      'DOI/link': link_entry
                }
    ws.cell(RefSL_entry + 1, 1).value = RefSL_entry
    ws.cell(RefSL_entry + 1, 2).value = p_w_entry
    ws.cell(RefSL_entry + 1, 3).value = citation_entry
    ws.cell(RefSL_entry + 1, 4).value = year_entry
    ws.cell(RefSL_entry + 1, 5).value = link_entry
    print('\n')
    print(new_entry, end='\n')
    save = str(input("save this new entry? [y/n]: "))
    if save == 'y':
        wb.save(xlsx_file_directory)
        print(f"\n[+]Reference saved. Check ref[{RefSL_entry}] in row {RefSL_entry + 1}")
    else:
        print("\n[-]Reference not saved!")
else:
    print('\n[-]Reference used before.')
    print('Check reference no. ',citation_column_list.index(citation_entry))
